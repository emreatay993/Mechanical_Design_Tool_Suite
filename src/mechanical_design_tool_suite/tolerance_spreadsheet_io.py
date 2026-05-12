"""CSV/XLSX import for vNext tolerance stackup projects."""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Any, Iterable

from .tolerance_catalog import ToleranceCatalog
from .tolerance_models import (
    DEFAULT_UNIT_SYSTEM,
    Flange,
    Joint,
    PathItem,
    SubJoint,
    ToleranceProject,
    sync_path_with_flanges,
)


BASE_REQUIRED_COLUMNS = {
    "joint",
    "sub_joint",
    "item_type",
    "item_name",
    "nominal_thickness",
}
SUPPORTED_ITEM_TYPES = {"flange", "custom", "catalog"}


def load_spreadsheet_project(
    path: str | Path,
    catalog: ToleranceCatalog | None = None,
) -> ToleranceProject:
    input_path = Path(path)
    if input_path.suffix.lower() == ".csv":
        rows = _read_csv_rows(input_path)
    elif input_path.suffix.lower() == ".xlsx":
        rows = _read_xlsx_rows(input_path)
    else:
        raise ValueError("Import supports .csv and .xlsx files.")
    return project_from_rows(rows, catalog or ToleranceCatalog.builtin())


def project_from_rows(
    rows: Iterable[dict[str, Any]],
    catalog: ToleranceCatalog | None = None,
) -> ToleranceProject:
    catalog = catalog or ToleranceCatalog.builtin()
    clean_rows = [_normalize_row(row) for row in rows if _row_has_values(row)]
    if not clean_rows:
        raise ValueError("Import table is empty.")
    _validate_headers(clean_rows[0])

    project_title = _first_value(clean_rows, "project_title") or "Imported Tolerance Project"
    unit_system = _first_value(clean_rows, "unit_system") or DEFAULT_UNIT_SYSTEM
    if unit_system != DEFAULT_UNIT_SYSTEM:
        raise ValueError("Only mm unit_system imports are supported.")

    project = ToleranceProject(title=project_title, unit_system=unit_system)
    joints_by_name: dict[str, Joint] = {}
    sub_joints_by_key: dict[tuple[str, str], SubJoint] = {}
    flange_specs_by_key: dict[tuple[str, str], tuple[float, float, float]] = {}

    for row_number, row in enumerate(clean_rows, start=2):
        joint_name = _required_text(row, "joint", row_number)
        sub_joint_name = _required_text(row, "sub_joint", row_number)
        item_type = _required_text(row, "item_type", row_number).lower()
        if item_type not in SUPPORTED_ITEM_TYPES:
            raise ValueError(
                f"Row {row_number}: item_type must be flange, custom, or catalog."
            )
        item_name = _required_text(row, "item_name", row_number)
        nominal = _required_float(row, "nominal_thickness", row_number)
        tolerance_minus, tolerance_plus = _row_tolerances(row, row_number)

        joint = joints_by_name.get(joint_name)
        if joint is None:
            joint = Joint(name=joint_name)
            project.joints.append(joint)
            joints_by_name[joint_name] = joint

        sub_joint_key = (joint_name, sub_joint_name)
        sub_joint = sub_joints_by_key.get(sub_joint_key)
        if sub_joint is None:
            sub_joint = SubJoint(name=sub_joint_name)
            joint.sub_joints.append(sub_joint)
            sub_joints_by_key[sub_joint_key] = sub_joint
        _apply_sub_joint_row_settings(sub_joint, row, catalog)

        if item_type == "flange":
            flange_key = (joint_name, item_name)
            spec = (nominal, tolerance_minus, tolerance_plus)
            existing = flange_specs_by_key.get(flange_key)
            if existing is not None and existing != spec:
                raise ValueError(
                    f"Row {row_number}: flange {item_name} has conflicting values."
                )
            if existing is None:
                flange_specs_by_key[flange_key] = spec
                joint.flanges.append(
                    Flange(
                        name=item_name,
                        nominal_thickness=nominal,
                        tolerance=max(tolerance_minus, tolerance_plus),
                        tolerance_minus=tolerance_minus,
                        tolerance_plus=tolerance_plus,
                    )
                )
            continue

        sub_joint.stackup_path.items.append(
            PathItem(
                source_type=item_type,
                source_id=str(row.get("source_id") or ""),
                name=item_name,
                nominal_thickness=nominal,
                tolerance=max(tolerance_minus, tolerance_plus),
                tolerance_minus=tolerance_minus,
                tolerance_plus=tolerance_plus,
                role=str(row.get("role") or item_type),
                include_in_stackup=_parse_bool(row.get("include_in_stackup"), True),
            )
        )

    if not project.joints:
        raise ValueError("Import did not create any joints.")
    for joint in project.joints:
        if not joint.sub_joints:
            joint.sub_joints.append(SubJoint(name=f"{joint.name}.1"))
        for sub_joint in joint.sub_joints:
            sync_path_with_flanges(joint, sub_joint)
    return project


def _read_csv_rows(path: Path) -> list[dict[str, Any]]:
    with path.open("r", newline="", encoding="utf-8-sig") as stream:
        reader = csv.DictReader(stream)
        return [dict(row) for row in reader]


def _read_xlsx_rows(path: Path) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency is declared.
        raise ValueError("XLSX import requires openpyxl.") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = (
            workbook["stackups"]
            if "stackups" in workbook.sheetnames
            else workbook[workbook.sheetnames[0]]
        )
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [_normalize_header(value) for value in rows[0]]
        output: list[dict[str, Any]] = []
        for values in rows[1:]:
            output.append(
                {
                    header: value
                    for header, value in zip(headers, values, strict=False)
                    if header
                }
            )
        return output
    finally:
        workbook.close()


def _normalize_row(row: dict[str, Any]) -> dict[str, Any]:
    return {
        _normalize_header(key): value.strip() if isinstance(value, str) else value
        for key, value in row.items()
        if _normalize_header(key)
    }


def _normalize_header(value: Any) -> str:
    return str(value or "").strip().lower().replace("-", "_").replace(" ", "_")


def _row_has_values(row: dict[str, Any]) -> bool:
    return any(value not in (None, "") for value in row.values())


def _validate_headers(row: dict[str, Any]) -> None:
    missing = sorted(BASE_REQUIRED_COLUMNS - set(row))
    has_symmetric = "tolerance" in row
    has_asymmetric = "tolerance_minus" in row and "tolerance_plus" in row
    if not has_symmetric and not has_asymmetric:
        missing.append("tolerance or tolerance_minus+tolerance_plus")
    if missing:
        raise ValueError(f"Import table is missing columns: {', '.join(missing)}.")


def _first_value(rows: list[dict[str, Any]], key: str) -> str:
    for row in rows:
        value = str(row.get(key) or "").strip()
        if value:
            return value
    return ""


def _required_text(row: dict[str, Any], key: str, row_number: int) -> str:
    value = str(row.get(key) or "").strip()
    if not value:
        raise ValueError(f"Row {row_number}: {key} is required.")
    return value


def _required_float(row: dict[str, Any], key: str, row_number: int) -> float:
    value = row.get(key)
    try:
        return float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"Row {row_number}: {key} must be numeric.") from exc


def _row_tolerances(row: dict[str, Any], row_number: int) -> tuple[float, float]:
    if row.get("tolerance_minus") not in (None, "") or row.get("tolerance_plus") not in (
        None,
        "",
    ):
        tolerance_minus = _required_float(row, "tolerance_minus", row_number)
        tolerance_plus = _required_float(row, "tolerance_plus", row_number)
    else:
        tolerance = _required_float(row, "tolerance", row_number)
        tolerance_minus = tolerance
        tolerance_plus = tolerance
    if tolerance_minus < 0.0 or tolerance_plus < 0.0:
        raise ValueError(f"Row {row_number}: tolerances must be non-negative.")
    return tolerance_minus, tolerance_plus


def _apply_sub_joint_row_settings(
    sub_joint: SubJoint,
    row: dict[str, Any],
    catalog: ToleranceCatalog,
) -> None:
    if row.get("bolt_size"):
        sub_joint.bolt_size_id = str(row["bolt_size"])
    if row.get("bolt_type"):
        sub_joint.bolt_type_id = str(row["bolt_type"])
    if row.get("bolt_length") not in (None, ""):
        sub_joint.selected_bolt_length = float(row["bolt_length"])
    if row.get("engagement_type"):
        sub_joint.stackup_path.engagement_type = str(row["engagement_type"]).lower()
    if row.get("engagement_part_id"):
        sub_joint.stackup_path.selected_engagement_part_id = str(row["engagement_part_id"])
    else:
        default = catalog.default_hardware(
            sub_joint.stackup_path.engagement_type,
            sub_joint.bolt_size_id,
        )
        if default:
            sub_joint.stackup_path.selected_engagement_part_id = default.id


def _parse_bool(value: Any, default: bool) -> bool:
    if value in (None, ""):
        return default
    if isinstance(value, bool):
        return value
    normalized = str(value).strip().lower()
    if normalized in {"1", "true", "yes", "y", "include"}:
        return True
    if normalized in {"0", "false", "no", "n", "exclude"}:
        return False
    raise ValueError(f"Invalid boolean value {value!r}.")
